import string

import openpyxl
from module.model import *
from module.GongziModel import *
from flask import Flask, render_template, request, redirect, url_for, session

from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm.exc import NoResultFound, MultipleResultsFound
from sqlalchemy import inspect


def random_code():
    # 定义包含字母和数字的字符集
    characters = string.ascii_letters + string.digits

    # 生成6位的随机字符串
    code = ''.join(random.choice(characters) for _ in range(13))

    return code.upper()

class PublicSysService:
    @staticmethod
    def add_record_to_table(model_class, **kwargs):
        record = model_class(**kwargs)
        if 'create_time' not in kwargs:
            record.create_time = datetime.now()  # 设置默认创建时间
        db.session.add(record)
        db.session.commit()
        return record

    @staticmethod
    def get_records_by_conditions(model_class, conditions):
        query = model_class.query
        for key, value in conditions.items():
            query = query.filter(getattr(model_class, key) == value)
        return query.all()

    @staticmethod
    def create_record(model_class, data):
        try:
            new_record = model_class(**data)  # 使用字典解包创建模型实例
            db.session.add(new_record)  # 将新记录添加到会话中
            db.session.commit()  # 提交会话以保存更改
            return new_record  # 返回新创建的记录
        except SQLAlchemyError as e:
            db.session.rollback()  # 发生错误时回滚会话
            return 0  # 返回错误信息和HTTP状态码500


    @staticmethod
    def get_all_records_from_table(model_class):
        return model_class.query.all()

    @staticmethod
    def update_record(model_class, id, **kwargs):
        ret = db.session.query(model_class).filter_by(id=id).first()
        if ret:
            for key, value in kwargs.items():
                setattr(ret, key, value)
            db.session.commit()
            return True
        return False

    @staticmethod
    def delete_record(model_class,id):
        ret = db.session.query(model_class).filter_by(id=id).first()
        if ret:
            db.session.delete(ret)
            db.session.commit()
            return True
        return False

    # @staticmethod
    # def add_and_updata_record(model_class, data):
    #     try:
    #         # 检查data中是否有'id'键，以区分是创建还是更新
    #         record_id = data.get('gonghao', None)
    #         if record_id is None:
    #             # 如果没有提供id，则认为是创建新记录
    #             new_record = model_class(**data)
    #             db.session.add(new_record)
    #         else:
    #             # 如果提供了id，则尝试更新现有记录
    #             try:
    #                 record = model_class.query.get(record_id)
    #                 if record is None:
    #                     record = model_class(**data)
    #                     if 'create_time' not in data:
    #                         record.create_time = datetime.now()  # 设置默认创建时间
    #                     db.session.add(record)
    #                     db.session.commit()
    #                     return record
    #
    #                 for key, value in data.items():
    #                     setattr(record, key, value)  # 更新记录属性
    #             except (NoResultFound, MultipleResultsFound) as e:
    #                 # 处理查询可能引发的异常（尽管get方法通常不会引发MultipleResultsFound）
    #                 raise ValueError(f"Error updating record with id {record_id}: {str(e)}")
    #
    #         db.session.commit()  # 提交会话以保存更改
    #         if record_id is None:
    #             return new_record  # 如果是创建，返回新创建的记录
    #         else:
    #             return record  # 如果是更新，返回更新后的记录
    #
    #     except SQLAlchemyError as e:
    #         db.session.rollback()  # 发生错误时回滚会话
    #         # 这里返回0可能不是最佳实践，通常应该抛出异常或返回更具体的错误信息
    #         # 但为了与原始代码保持一致，这里保留返回0
    #         # 在实际应用中，考虑使用日志记录错误，并向调用者抛出一个更具体的异常
    #         return 0
    @staticmethod
    def add_and_update_record(model_class, data):
        # try:
        for key, value in data.items():
            if value == "None":
                data[key] = ""
        if "id" not in data:
            record = None
            new_record = model_class(**data)
            if 'create_time' not in data:
                new_record.create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 设置默认创建时间
            db.session.add(new_record)
        else:
            # 如果提供了gonghao，则尝试更新现有记录
            id = data.get('id',None)
            if id:
                record = model_class.query.filter_by(id=id).first()
                if record is None:
                    # 如果记录不存在，则创建新记录
                    new_record = model_class(**data)
                    if 'create_time' not in data:
                        new_record.create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 设置默认创建时间
                    db.session.add(new_record)
                else:
                    # 如果记录存在，则更新该记录
                    for key, value in data.items():
                        setattr(record, key, value)  # 更新记录属性
            else:
                record = None
                data.pop("id")
                new_record = model_class(**data)
                if 'create_time' not in data:
                    new_record.create_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 设置默认创建时间
                db.session.add(new_record)
        db.session.commit()
        return record if record else new_record
        # except Exception as e:
        #     # 处理可能发生的任何异常
        #     db.session.rollback()
        #     raise ValueError(f"Error adding/updating record: {str(e)}")
    @staticmethod
    def get_user_info(username):
        user = db.session.query(User).filter_by(username=username).first()
        return user

    @staticmethod
    def get_user_list():
        return db.session.query(User).all()

    @staticmethod
    def set_user_role(username, role):
        user = db.session.query(User).filter_by(username=username).first()
        user.role = role
        db.session.add(user)
        db.session.commit()
        return user


    @staticmethod
    def get_record_list(position):
        if position:
            return db.session.query(Record).filter_by(position=position).order_by(Record.date.desc()).all()
        return db.session.query(Record).order_by(Record.date.desc()).all()

    @staticmethod
    def excel_import(model_class,filename, sheet_name=None):
        # try:
        obj = openpyxl.load_workbook(f"static/excel/{filename}")
        sheet = obj.active if sheet_name is None else obj[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        all_row_data = []
        first_row_values = []
        # 遍历第一行的所有单元格
        for cell in sheet[1]:
            first_row_values.append(cell.value)
        for i in range(2, max_row + 1):  # 假设第一行是标题行，从第二行开始读取数据
            row_data = [str(sheet.cell(i, col_idx).value) for col_idx in range(1, max_column + 1)]

            # 验证数据完整性
            if not all(row_data):
                return -1, f"第{i}行数据不完整"

            all_row_data.append(row_data)
        # 插入数据
        for row_data in all_row_data:
            row_dict = dict(zip(first_row_values, row_data))
            add_and_updata(model_class,row_dict)

        return 0, f"导入完成,共{len(all_row_data)}条"
        #
        # except Exception as e:
        #     return -1, str(e)
    @staticmethod
    def row_data_to_dict(model_class, row_data):
        # 获取模型的列信息
        mapper = inspect(model_class)
        columns = {c.key: None for c in mapper.columns}
        len_columns = len(columns)-1
        # 假设row_data是一个列表或元组，其顺序与model_class中的字段顺序相对应

        if len(row_data) != len_columns:
            raise ValueError("数据不匹配")

            # 将row_data转换为字典，键为字段名，值为对应的数据
        result = {}
        for i, value in enumerate(row_data):
            column_name = list(columns.keys())[i]  # 注意：这里依赖于字段顺序，通常不推荐
            result[column_name] = value

        return result
    @staticmethod
    def get_setting():
        s = db.session.query(Setting).first()
        if not s:
            s = Setting()
        return s


app = Flask(__name__)

HOST = "127.0.0.1"  # 数据库地址
PORT = 3306
USERNAME = "root"   # 数据库账号
PASSWORD = "123456"  # 数据库密码
DATABASE = "gongzi"  # 数据库名
app.config['SQLALCHEMY_DATABASE_URI'] = "mysql+pymysql://{}:{}@{}:{}/{}".format(USERNAME, PASSWORD, HOST, PORT, DATABASE)

app.config["SECRET_KEY"] = "JKXOSDIFHL@ksljfflaj@sdfasdf"
db.init_app(app)

def add_new_neitui(new_neitui_data):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        new_neitui_record = SysService.add_record_to_table(Neitui, **new_neitui_data)
        return new_neitui_record

def get_all_records(model_class):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        new_neitui_record = PublicSysService.get_all_records_from_table(model_class)

        return new_neitui_record

def del_records(model_class,id):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        ret = PublicSysService.delete_record(model_class,id)

        return ret

def create_records(model_class, data):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        ret = PublicSysService.create_record(model_class, data)

        return ret

def update_records(model_class, id, **kwargs):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        ret = PublicSysService.update_record(model_class, id, **kwargs)
        return ret

def add_and_updata(model_class, data):
    with app.app_context():
        # 在这里，应用上下文已经被设置
        ret = PublicSysService.add_and_update_record(model_class, data)

        return ret


def get_create_data_from_model(model, request_data):
    """
    根据模型的字段和request_data中的数据，动态构建create_data字典。
    """
    create_data = {}
    # 获取模型的所有字段名（除了主键，假设主键是自增的或者由其他方式生成）
    model_fields = [field.name for field in model.__table__.columns ]
    # 遍历字段名，从request_data中获取对应的值
    for field in model_fields:
        value = request_data.get(field, "")
        if value is not None:
            create_data[field] = value

            # 对于特殊字段，如时间戳，进行特殊处理
    if 'create_time' in model_fields:
        create_data['create_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 自动设置当前时间
    return create_data