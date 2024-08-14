import string

import openpyxl
from module.model import *


def random_code():
    # 定义包含字母和数字的字符集
    characters = string.ascii_letters + string.digits

    # 生成6位的随机字符串
    code = ''.join(random.choice(characters) for _ in range(13))

    return code.upper()


class SysService:

    @staticmethod
    def get_user_info(username):
        user = db.session.query(User).filter_by(username=username).first()
        return user

    @staticmethod
    def get_user_list():
        return db.session.query(User).all()

    @staticmethod
    def set_user_role(username, role):
        user = db.session.query(User).filter_by(username=username).first()
        user.role = role
        db.session.add(user)
        db.session.commit()
        return user

    @staticmethod
    def add_record(record_id, position, date, jsl, ysl, sw, temp, sls, dbcxwy, wz):
        r = Record()
        if record_id != "-1":
            r = db.session.query(Record).filter_by(id=record_id).first()
            if not r:
                return -1, "未查找到该记录"
        r.position = position
        r.jsl = jsl
        r.ysl = ysl
        r.sw = sw
        r.temp = temp
        r.sls = sls
        r.dbcxwy = dbcxwy
        r.wz = wz
        r.date = date
        db.session.add(r)
        db.session.commit()

    @staticmethod
    def delete_record(record_id):
        ret = db.session.query(Record).filter_by(id=record_id).first()
        if ret:
            db.session.delete(ret)
            db.session.commit()
            return True
        return False

    @staticmethod
    def get_record_list(position):
        if position:
            return db.session.query(Record).filter_by(position=position).order_by(Record.date.desc()).all()
        return db.session.query(Record).order_by(Record.date.desc()).all()

    @staticmethod
    def set_flag(record: Record, setting: Setting):
        fields = ['jsl', 'ysl', 'sw', 'temp', 'sls', 'dbcxwy', 'wz']
        for filed in fields:
            if getattr(setting, filed) and getattr(setting, filed)['value']:
                if getattr(setting, filed)['action'] == 'lt':
                    if float(getattr(record, filed)) <= float(getattr(setting, filed)['value']):
                        return True
                else:
                    if float(getattr(record, filed)) >= float(getattr(setting, filed)['value']):
                        return True


    @staticmethod
    def update_setting(data):
        s = db.session.query(Setting).first()
        if not s:
            s = Setting()

        s.jsl = data.get("jsl")
        s.ysl = data.get("ysl")
        s.sw = data.get("sw")
        s.temp = data.get("temp")
        s.sls = data.get("sls")
        s.dbcxwy = data.get("dbcxwy")
        s.wz = data.get("wz")
        db.session.add(s)
        db.session.commit()

    @staticmethod
    def excel_import(filename):
        obj = openpyxl.load_workbook(f"static/excel/{filename}")
        sheet = obj.active
        max_row = sheet.max_row
        all_row_data = []
        for i in range(2, max_row + 1):
            date = sheet.cell(i, 2).value
            if not isinstance(date, datetime):
                return -1, f"第{i}行日期格式错误"
            position = str(sheet.cell(i, 1).value)
            jsl = str(sheet.cell(i, 3).value)
            ysl = str(sheet.cell(i, 4).value)
            sw = str(sheet.cell(i, 5).value)
            temp = str(sheet.cell(i, 6).value)
            sls = str(sheet.cell(i, 7).value)
            dbcxwy = str(sheet.cell(i, 8).value)
            wz = str(sheet.cell(i, 9).value)
            if not all([position, jsl, ysl, sw, temp, sls, dbcxwy, wz]):
                return -1, f"第{i}行数据不完整"
            if not all([x.isdigit() for x in [jsl, ysl, sw, temp, sls, dbcxwy, wz]]):
                return -1, f"第{i}行数据格式错误"
            all_row_data.append([position, date, jsl, ysl, sw, temp, sls, dbcxwy, wz])
        # 插入数据
        for row_data in all_row_data:
            SysService.add_record("-1", *row_data)

        return 0, f"导入完成,共{len(all_row_data)}条"

    @staticmethod
    def get_setting():
        s = db.session.query(Setting).first()
        if not s:
            s = Setting()
        return s
